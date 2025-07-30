from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from io import BytesIO
import pandas as pd
from typing import List, Optional


class ExcelProcessor:
    @staticmethod
    def load_workbook_from_bytes(excel_bytes: bytes) -> Workbook:
        """
        Load an Excel workbook from a byte stream.

        Args:
            excel_bytes (bytes): Raw Excel content as bytes.

        Returns:
            Workbook: An openpyxl Workbook object.

        Raises:
            ValueError: If no bytes are provided.
            RuntimeError: If loading the workbook fails.
        """
        if not excel_bytes:
            raise ValueError("No Excel bytes provided.")

        try:
            return load_workbook(BytesIO(excel_bytes), data_only=True)
        except Exception as e:
            raise RuntimeError(f"Failed to load Excel file: {e}") from e

    @staticmethod
    def extract_sheet_names(workbook: Workbook) -> List[str]:
        """
        Get the list of sheet names from a workbook.

        Args:
            workbook (Workbook): The loaded Excel workbook.

        Returns:
            List[str]: List of sheet names.
        """
        return workbook.sheetnames

    @staticmethod
    def get_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
        """
        Retrieve a worksheet by name.

        Args:
            workbook (Workbook): The Excel workbook.
            sheet_name (str): The name of the sheet to retrieve.

        Returns:
            Worksheet: The specified worksheet.

        Raises:
            RuntimeError: If the sheet cannot be retrieved.
        """
        try:
            return workbook[sheet_name]
        except Exception as e:
            raise RuntimeError(f"Failed to access sheet '{sheet_name}': {e}") from e

    @staticmethod
    def extract_sheet_data(sheet: Worksheet) -> List[List[Optional[str]]]:
        """
        Extract all cell values from a worksheet as a 2D list.

        Args:
            sheet (Worksheet): The worksheet object.

        Returns:
            List[List[Optional[str]]]: Sheet data as rows of cell values.
        """
        return [[cell.value for cell in row] for row in sheet.iter_rows()]

    @staticmethod
    def convert_sheet_to_dataframe(sheet: Worksheet) -> pd.DataFrame:
        """
        Convert a worksheet to a pandas DataFrame using the first row as headers.

        Args:
            sheet (Worksheet): The worksheet to convert.

        Returns:
            pd.DataFrame: The sheet data as a DataFrame.

        Raises:
            RuntimeError: If conversion fails or headers are missing.
        """
        try:
            data = ExcelProcessor.extract_sheet_data(sheet)
            if not data or not data[0]:
                raise RuntimeError("Sheet is empty or has no header row.")

            headers = data[0]
            rows = data[1:]
            return pd.DataFrame(rows, columns=headers)

        except Exception as e:
            raise RuntimeError(f"Failed to convert sheet to DataFrame: {e}") from e
